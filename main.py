import os
import threading
from tkinter import messagebox

from openpyxl import load_workbook
from datetime import date

from tkinter import *

root = Tk()
root.geometry('250x100')
root.iconbitmap(r'MC.ico')
root.resizable(False, False)

__salary = StringVar()

wb = load_workbook('MoneyControl.xlsx')
wb.iso_dates = True
ws = wb.active

MAX_ROW = ws.max_row

FORMULAS = [
    date.today().strftime('%d-%m-%Y'),                     # date
    f'=(A{MAX_ROW + 1}-A{MAX_ROW})',                       # salary increment 
    f'=ROUND((C{MAX_ROW + 1}/A{MAX_ROW})*100,0)',          # % increase
    '',
    f'=ROUNDDOWN((F{MAX_ROW}+(C{MAX_ROW + 1}*S3/100)),0)', # Expenses
    f'=ROUNDDOWN((G{MAX_ROW}+(C{MAX_ROW + 1}*T3/100)),0)', # Want
    f'=ROUNDUP((H{MAX_ROW}+(C{MAX_ROW + 1}*U3/100)),0)',   # Invest
    '',
    f'=ROUNDUP(H{MAX_ROW + 1}*J3/100,0)',                  # Equity
    f'=ROUNDDOWN(H{MAX_ROW + 1}*K3/100,0)',                # Debt
    '',
    f'=ROUNDUP(J{MAX_ROW + 1}*M3/100,0)',                  # index
    f'=ROUNDUP(J{MAX_ROW + 1}*N3/100,0)',                  # flexi
    f'=ROUNDDOWN(J{MAX_ROW + 1}*O3/100,0)',                # mid
    f'=ROUNDDOWN(J{MAX_ROW + 1}*P3/100,0)',                # small
]


def submit():
    inputData = __salary.get()
    if not inputData.isdigit():
        messagebox.showerror("Error", "Please enter valid in-hand salary")
        entry.delete(0, 'end')
        return

    dataToBeAdded = [int(inputData), ]
    dataToBeAdded = dataToBeAdded + FORMULAS

    print(dataToBeAdded)

    ws.append(dataToBeAdded)

    wb.save('MoneyControl.xlsx')

    wb.close()
    root.destroy()

    os.startfile('MoneyControl.xlsx')


Label(root, text="Enter in-hand salary").pack(side='top')
entry = Entry(root, textvariable=__salary, justify='center', bd=3, font='18')
entry.pack(expand=True, fill='both', side='top')
Button(root, text='Submit', command=submit).pack()

root.mainloop()
